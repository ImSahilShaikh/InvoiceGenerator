# Using openpyxl to operate on excel file
import openpyxl

# Using reportlab to create a pdf file
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont

# PIL for operating on images
#from PIL import Image

# importing sys for reading commandline arguments
import sys

# convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))

# import the sheet from the excel file
wb = openpyxl.load_workbook(filename = 'Book1.xlsx')
sheet = wb["hey"]

# Uncomment to add logo 
# add image height and width    
# im = Image.open('logo.jpg')
# width, height = im.size
# ratio = width/height
# image_width = 400
# image_height = int(image_width / ratio)

# Page information
page_width = 2156
page_height = 3050

# Invoice variables
company_name = 'InvoiceX Inc.'
payment_terms = 'x'
contact_info = 'x'
margin = 100
month_year = 'August 2019'

# def function
def create_invoice():
    row_number=sys.argv[1]
    i = int(row_number)
    # Reading values from excel file
    customer = sheet.cell(row=i, column=3).value
    invoice_number = sheet.cell(row=i, column=1).value
    invoice_date = sheet.cell(row=i, column=8).value
    due_date = sheet.cell(row=i, column=9).value
    description = sheet.cell(row=i, column=4).value.lower()
    amount_excl_vat = sheet.cell(row=i, column=5).value
    vat = sheet.cell(row=i, column=6).value
    total_amount = sheet.cell(row=i, column=7).value

    # Creating a pdf file and setting a naming convention
    c = canvas.Canvas(str(invoice_number) + '_' + str(customer) + '.pdf')
    c.setPageSize((page_width, page_height))

    # Uncomment to add the image logo
    # c.drawInlineImage("logo.jpg", page_width - image_width - margin,
    #                   page_height - image_height - margin,
    #                   image_width, image_height)

    # Invoice information
    # c.setFont('Arial', 80)
    # text = 'INVOICE'
    # text_width = stringWidth(text, 'Arial', 80)
    # c.drawString((page_width-text_width)/2, page_height -
    # - margin, text)
    y = page_height - margin*4
    x = 2*margin
    x2 = x + 550

    c.setFont("Arial",80)
    c.drawString(x,y,"INVOICE")
    y -=margin

    c.setFont('Arial', 45)
    c.drawString(x, y, 'Issued by: ')
    c.drawString(x2, y, company_name)
    y -= margin

    c.drawString(x, y, 'Issued to: ')
    c.drawString(x2, y, customer)
    y -= margin

    c.drawString(x, y, 'Invoice number: ')
    c.drawString(x2, y, str(invoice_number))
    y -= margin

    c.drawString(x, y, 'Invoice date: ')
    c.drawString(x2, y, invoice_date)
    y -= margin

    c.drawString(x, y, 'Due date: ')
    c.drawString(x2, y, due_date)
    y -= margin * 2

    c.drawString(x,y,'Invoice issued for performed '+ description + ' for ' + month_year)
    y -= margin *2
        
    c.drawString(x,y, 'Amount excluding VAT: ')
    c.drawString(x2,y, 'EUR ' + str(amount_excl_vat))
    y -= margin
        
    c.drawString(x,y,'Value added tax: ')
    c.drawString(x2,y, 'EUR ' + str(vat))
    y-= margin
        
    c.drawString(x,y,'Total amount: ')
    c.drawString(x2,y,'EUR ' + str(total_amount))
    y -= margin*3
               
    c.drawString(x,y,'If paid within 10 days, 2% of discount is granted.')
    y -= margin
    c.drawString(x,y,'In case of any questions, contact info@invoicex.com')
    y -= margin

    # Saving the pdf file
    c.save()

create_invoice()
